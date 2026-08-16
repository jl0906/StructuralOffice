"""Excel import and export for StructuralOffice accounting data."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .accounting import (
    DIRECTION_PAYABLE,
    DIRECTION_RECEIVABLE,
    INVOICE_STATUS_CANCELLED,
    INVOICE_STATUS_OPEN,
    INVOICE_STATUS_PAID,
    cents_to_amount,
    validate_invoice,
)
from .models import StructuralOfficeValidationError

SHEET_NAME = "Accounting"
MAX_FILE_SIZE = 5 * 1024 * 1024
MAX_UNCOMPRESSED_SIZE = 25 * 1024 * 1024
MAX_ROWS = 5000
TEMPLATE_PATH = Path(__file__).parent / "assets" / "StructuralOffice-Accounting-Template.xlsx"

HEADERS = [
    "ID",
    "Type",
    "Contact",
    "Invoice Number",
    "Invoice Date",
    "Due Date",
    "Net Amount",
    "Tax Amount",
    "Gross Amount",
    "Currency",
    "Status",
    "Paid On",
    "Dunning Level",
    "Payment Reminders (Days)",
    "Dunning Periods (Days)",
    "Note",
    "Last Modified",
]

DIRECTION_LABELS = {
    DIRECTION_PAYABLE: "Payable",
    DIRECTION_RECEIVABLE: "Receivable",
}
STATUS_LABELS = {
    INVOICE_STATUS_OPEN: "Open",
    INVOICE_STATUS_PAID: "Paid",
    INVOICE_STATUS_CANCELLED: "Cancelled",
}


def load_template_xlsx() -> bytes:
    """Load the visually verified bundled workbook template."""
    return TEMPLATE_PATH.read_bytes()


def _safe_workbook(content: bytes):
    """Load a size-limited XLSX workbook."""
    if not content or len(content) > MAX_FILE_SIZE:
        raise StructuralOfficeValidationError("Excel file is empty or larger than 5 MB")
    try:
        with ZipFile(BytesIO(content)) as archive:
            infos = archive.infolist()
            if len(infos) > 250 or sum(item.file_size for item in infos) > MAX_UNCOMPRESSED_SIZE:
                raise StructuralOfficeValidationError("Uncompressed Excel file is too large")
    except BadZipFile as err:
        raise StructuralOfficeValidationError("File is not a valid XLSX workbook") from err
    try:
        return load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as err:
        raise StructuralOfficeValidationError("Excel file could not be read") from err


def _normalized_header(value: Any) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def _as_iso_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return text


def _direction(value: Any) -> str:
    normalized = _normalized_header(value)
    mapping = {
        "eingangsrechnung": DIRECTION_PAYABLE,
        "eingang": DIRECTION_PAYABLE,
        "payable": DIRECTION_PAYABLE,
        "ausgangsrechnung": DIRECTION_RECEIVABLE,
        "ausgang": DIRECTION_RECEIVABLE,
        "receivable": DIRECTION_RECEIVABLE,
    }
    return mapping.get(normalized, normalized)


def _status(value: Any) -> str:
    normalized = _normalized_header(value)
    mapping = {
        "": INVOICE_STATUS_OPEN,
        "offen": INVOICE_STATUS_OPEN,
        "open": INVOICE_STATUS_OPEN,
        "bezahlt": INVOICE_STATUS_PAID,
        "paid": INVOICE_STATUS_PAID,
        "storniert": INVOICE_STATUS_CANCELLED,
        "cancelled": INVOICE_STATUS_CANCELLED,
        "canceled": INVOICE_STATUS_CANCELLED,
    }
    return mapping.get(normalized, normalized)


def parse_invoices_xlsx(content: bytes) -> dict[str, Any]:
    """Parse and validate the StructuralOffice accounting worksheet."""
    workbook = _safe_workbook(content)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as err:
        raise StructuralOfficeValidationError("Excel worksheet is empty") from err

    header_map = {_normalized_header(value): index for index, value in enumerate(raw_headers)}
    required = {
        "Type": ("type", "typ"),
        "Contact": ("contact", "kontakt"),
        "Invoice Number": ("invoice number", "rechnungsnummer"),
        "Invoice Date": ("invoice date", "rechnungsdatum"),
        "Due Date": ("due date", "fälligkeitsdatum"),
    }
    missing = [
        label
        for label, aliases in required.items()
        if not any(alias in header_map for alias in aliases)
    ]
    if missing:
        raise StructuralOfficeValidationError(f"Required columns are missing: {', '.join(missing)}")

    def cell(row: tuple[Any, ...], *names: str) -> Any:
        index = next(
            (
                header_map[_normalized_header(name)]
                for name in names
                if _normalized_header(name) in header_map
            ),
            None,
        )
        return row[index] if index is not None and index < len(row) else None

    records: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    seen_business_keys: set[tuple[str, str, str]] = set()
    for row_number, row in enumerate(rows, start=2):
        if row_number > MAX_ROWS + 1:
            errors.append(
                {
                    "row": row_number,
                    "message": f"A maximum of {MAX_ROWS} records is allowed",
                }
            )
            break
        identity_values = (
            cell(row, "Contact", "Kontakt"),
            cell(row, "Invoice Number", "Rechnungsnummer"),
            cell(row, "Invoice Date", "Rechnungsdatum"),
            cell(row, "Due Date", "Fälligkeitsdatum"),
        )
        if not any(value not in (None, "") for value in identity_values):
            continue
        raw = {
            "id": cell(row, "ID"),
            "direction": _direction(cell(row, "Type", "Typ")),
            "contact": cell(row, "Contact", "Kontakt"),
            "invoice_number": cell(row, "Invoice Number", "Rechnungsnummer"),
            "invoice_date": _as_iso_date(cell(row, "Invoice Date", "Rechnungsdatum")),
            "due_date": _as_iso_date(cell(row, "Due Date", "Fälligkeitsdatum")),
            "net_amount": cell(row, "Net Amount", "Nettobetrag"),
            "tax_amount": cell(row, "Tax Amount", "Steuerbetrag"),
            "gross_amount": cell(row, "Gross Amount", "Bruttobetrag"),
            "currency": cell(row, "Currency", "Währung") or "EUR",
            "status": _status(cell(row, "Status")),
            "paid_date": _as_iso_date(cell(row, "Paid On", "Bezahlt am")),
            "dunning_level": cell(row, "Dunning Level", "Mahnstufe") or 0,
            "payment_reminder_offsets": cell(
                row, "Payment Reminders (Days)", "Zahlungserinnerungen (Tage)"
            ),
            "dunning_offsets": cell(row, "Dunning Periods (Days)", "Mahnfristen (Tage)"),
            "note": cell(row, "Note", "Notiz"),
        }
        try:
            invoice = validate_invoice(raw)
            if invoice["id"] in seen_ids:
                errors.append({"row": row_number, "message": "ID is duplicated in the file"})
                continue
            seen_ids.add(invoice["id"])
            business_key = (
                invoice["direction"],
                invoice["contact"].casefold(),
                invoice["invoice_number"].casefold(),
            )
            if business_key in seen_business_keys:
                warnings.append(
                    {
                        "row": row_number,
                        "message": "Possible duplicate invoice number for this contact",
                    }
                )
            seen_business_keys.add(business_key)
            records.append(invoice)
        except (StructuralOfficeValidationError, TypeError, ValueError) as err:
            errors.append({"row": row_number, "message": str(err)})
    workbook.close()
    return {"records": records, "errors": errors, "warnings": warnings}


def export_invoices_xlsx(invoices: list[dict[str, Any]]) -> bytes:
    """Build a formatted StructuralOffice XLSX export."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.append(HEADERS)

    for invoice in invoices:
        sheet.append(
            [
                invoice["id"],
                DIRECTION_LABELS[invoice["direction"]],
                invoice["contact"],
                invoice["invoice_number"],
                date.fromisoformat(invoice["invoice_date"]),
                date.fromisoformat(invoice["due_date"]),
                float(cents_to_amount(invoice["net_cents"])),
                float(cents_to_amount(invoice["tax_cents"])),
                float(cents_to_amount(invoice["gross_cents"])),
                invoice["currency"],
                STATUS_LABELS[invoice["status"]],
                date.fromisoformat(invoice["paid_date"]) if invoice.get("paid_date") else None,
                invoice["dunning_level"],
                ", ".join(str(item) for item in invoice["payment_reminder_offsets"]),
                ", ".join(str(item) for item in invoice["dunning_offsets"]),
                invoice["note"],
                invoice["updated_at"],
            ]
        )

    if not invoices:
        sheet.append(
            [
                None,
                "Payable",
                "",
                "",
                None,
                None,
                0,
                0,
                0,
                "EUR",
                "Open",
                None,
                0,
                "-7, -1, 0",
                "3, 10, 20",
                "",
                "",
            ]
        )

    header_fill = PatternFill("solid", fgColor="0F766E")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34

    widths = [20, 20, 25, 20, 16, 16, 16, 16, 16, 11, 13, 16, 12, 25, 22, 35, 25]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = width
    for row in range(2, sheet.max_row + 1):
        for column in (5, 6, 12):
            sheet.cell(row, column).number_format = "yyyy-mm-dd"
        for column in (7, 8, 9):
            sheet.cell(row, column).number_format = "#,##0.00 [$€-407]"

    type_validation = DataValidation(
        type="list", formula1='"Payable,Receivable"', allow_blank=False
    )
    status_validation = DataValidation(
        type="list", formula1='"Open,Paid,Cancelled"', allow_blank=False
    )
    sheet.add_data_validation(type_validation)
    sheet.add_data_validation(status_validation)
    type_validation.add("B2:B5001")
    status_validation.add("K2:K5001")
    sheet.conditional_formatting.add(
        "A2:Q5001",
        FormulaRule(
            formula=['AND($K2="Open",$F2<TODAY(),$F2<>"")'],
            fill=PatternFill("solid", fgColor="FEE2E2"),
        ),
    )
    sheet.auto_filter.ref = f"A1:Q{max(2, sheet.max_row)}"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
